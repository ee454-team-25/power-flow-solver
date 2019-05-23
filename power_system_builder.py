"""A module that builds power system objects from input files.

Currently supported file formats:

    1. Excel
"""

import openpyxl
import power_system

# Builder defaults.
FLAT_START_VOLTAGE = 1 + 0j
DEFAULT_BUS_DATA_WORKSHEET_NAME = 'Bus data'
DEFAULT_LINE_DATA_WORKSHEET_NAME = 'Line data'
DEFAULT_POWER_BASE = 100


class PowerSystemBuilder:
    def build_buses(self):
        raise NotImplementedError()

    def build_lines(self):
        raise NotImplementedError()

    def build_system(self):
        """Builds a power system."""
        return power_system.PowerSystem(self.build_buses(), self.build_lines())


class ExcelPowerSystemBuilder(PowerSystemBuilder):
    """A power system builder that reads data from Excel files.

    The input is expected to be a spreadsheet with two worksheets: one specifying buses and another specifying lines.
    By default the worksheets are expected to be called "Bus data" and "Line data," but these values may be overridden.

    The bus data worksheet is expected to have a header row and data in the following format:

        1. Bus number
        2. Active power consumed in MW
        3. Reactive power consumed in Mvar
        4. Active power generated in MW
        5. Voltage at the bus in per-unit

    The line data worksheet is expected to have a header row and data in the following format:

        1. The source bus number
        2. The destination bus number
        3. The distributed resistance in per-unit
        4. The distributed reactance in per-unit
        5. The shunt susceptance in per-unit
        6. The maximum power rating of the line in MVA
    """

    def __init__(self, filename, bus_data_worksheet_name=DEFAULT_BUS_DATA_WORKSHEET_NAME,
                 line_data_worksheet_name=DEFAULT_LINE_DATA_WORKSHEET_NAME, start_voltage=FLAT_START_VOLTAGE,
                 power_base=DEFAULT_POWER_BASE):
        """Initializes the power system builder.

        Args:
            filename: The Excel workbook filename.
            bus_data_worksheet_name: The name of the worksheet containing bus data.
            line_data_worksheet_name: The name of the worksheet containing line data.
            start_voltage: The start voltage for PQ buses.
            power_base: The power base in MVA.
        """
        self._workbook = openpyxl.load_workbook(filename, read_only=True)
        self._bus_data_worksheet = self._workbook[bus_data_worksheet_name]
        self._line_data_worksheet = self._workbook[line_data_worksheet_name]
        self._start_voltage = start_voltage
        self._power_base = power_base

    def build_buses(self):
        """Builds a list of buses in the system."""
        result = []
        for row in self._bus_data_worksheet.iter_rows(min_row=2):
            bus_number = row[0].value
            if not bus_number:
                break

            p_load = (row[1].value or 0) / self._power_base
            q_load = (row[2].value or 0) / self._power_base
            p_generator = (row[3].value or 0) / self._power_base
            p_voltage = row[4].value or self._start_voltage
            result.append(power_system.Bus(bus_number, p_load, q_load, p_generator, p_voltage))

        return result

    def build_lines(self):
        """Builds a list of lines in the system."""
        result = []
        for row in self._line_data_worksheet.iter_rows(min_row=2):
            source_bus_number = row[0].value
            destination_bus_number = row[1].value
            if not source_bus_number or not destination_bus_number:
                break

            r_distributed = row[2].value or 0
            x_distributed = row[3].value or 0
            z_distributed = r_distributed + 1j * x_distributed
            y_shunt = 1j * row[4].value / 2 or 0j
            max_power = row[5].value

            result.append(
                power_system.Line(source_bus_number, destination_bus_number, z_distributed, y_shunt, max_power))

        return result
